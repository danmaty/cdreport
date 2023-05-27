import math
from time import strftime as stt
import streamlit as st
import polars as pl
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from io import BytesIO
# from deta import Deta
import zulip
import os
import warnings
warnings.simplefilter(action='ignore')


def make_figs(file):
    ##############################################
    #   Make df with figs and list of names
    ##############################################
    activities_to_keep = ['Pack complete', 'Pick from location', 'Receiving', 'Putaway putdown', 'Move dropoff']

    try:
        df = pl.read_csv(file, infer_schema_length=0, sep=',').lazy() \
            .filter(pl.col('Movement Type').is_in(activities_to_keep)) \
            .groupby(pl.col('Username')) \
            .agg([
            (pl.col('Movement Type')).count().alias('Figures'),
            (pl.col('Movement Type').str.contains('Pick from location')).sum().alias('Picks'),
            (pl.col('Movement Type').str.contains('Pack complete')).sum().alias('Packs'),
            (pl.col('Movement Type').str.contains('Putaway putdown')).sum().alias('Putaway'),
            (pl.col('Movement Type').str.contains('Move dropoff')).sum().alias('Moves'),
            (pl.col('Movement Type').str.contains('Receiving')).sum().alias('Rcv'),
            (pl.col('Movement Date').str.contains(' 06:')).sum().alias('7am'),
            (pl.col('Movement Date').str.contains(' 07:')).sum().alias('8am'),
            (pl.col('Movement Date').str.contains(' 08:')).sum().alias('9am'),
            (pl.col('Movement Date').str.contains(' 09:')).sum().alias('10am'),
            (pl.col('Movement Date').str.contains(' 10:')).sum().alias('11am'),
            (pl.col('Movement Date').str.contains(' 11:')).sum().alias('12pm'),
            (pl.col('Movement Date').str.contains(' 12:')).sum().alias('13pm'),
            (pl.col('Movement Date').str.contains(' 13:')).sum().alias('14pm'),
            (pl.col('Movement Date').str.contains(' 14:')).sum().alias('15pm'),
            (pl.col('Movement Date').str.contains(' 15:')).sum().alias('16pm'),
            (pl.col('Movement Date').str.contains(' 16:')).sum().alias('17pm'),
            (pl.col('Movement Date').str.contains(' 17:')).sum().alias('18pm'),
            (pl.col('Movement Date').str.contains(' 18:')).sum().alias('19pm'),
            (pl.col('Movement Date').str.contains(' 19:')).sum().alias('20pm'),
            (pl.col('Movement Date').str.contains(' 20:')).sum().alias('21pm'),
            (pl.col('Movement Date').str.contains(' 21:')).sum().alias('22pm'),
            (pl.col('Movement Date').str.contains(' 22:')).sum().alias('23pm')
        ]) \
            .filter(pl.col('Figures') > 0) \
            .sort(pl.col('Username'))

        out = df.collect().to_pandas()
        out.rename({'Username': 'Name'}, axis=1, inplace=True)

        return out, out['Name'].to_list()

    except Exception as e:
        print('make_figs', e)


def dt_row_func(row):
    if row['Name'] == row['Name2']:
        return row['Time'] - row['Time2']
    else:
        return None


def make_dt(file, downtime, keep):
    ##############################################
    #   Make df with dt / time data
    ##############################################
    try:
        df = pd.read_csv(file, sep=',', on_bad_lines='skip')
        df = df[['Username', 'Movement Date']]
        df = df[df['Username'].isin(keep)]

        ##############################################
        #   DT
        ##############################################
        df.rename({'Movement Date': 'Time', 'Username': 'Name'}, axis=1, inplace=True)
        df.Time = pd.to_datetime(df.Time)
        df = df.sort_values(['Name', 'Time'])
        df['Name2'] = df['Name'].shift(1)
        df['Time2'] = df['Time'].shift(1)
        df['Downtime'] = df.apply(lambda row: dt_row_func(row), axis=1)
        del df['Name2']
        del df['Time2']
        df['Downtime'] = df['Downtime'].shift(-1)

        ##############################################
        #   Make dict for start
        ##############################################
        df_start = df.groupby('Name')['Time'].min()
        df_start = df_start.reset_index()
        start_dict = dict(zip(df_start['Name'], df_start['Time']))

        ##############################################
        #   Make dict for finish
        ##############################################
        df_finish = df.groupby('Name')['Time'].max()
        df_finish = df_finish.reset_index()
        finish_dict = dict(zip(df_finish['Name'], df_finish['Time']))

        ##############################################
        #   Make dict for time (finish - start)
        ##############################################
        time_dict = {}
        for k in df['Name'].tolist():
            time_dict.update({k: finish_dict.get(k) - start_dict.get(k)})

        ##############################################
        #   Downtime filter
        ##############################################
        df = df[df['Downtime'] > downtime]

        ##############################################
        #   Make dict for dt count
        ##############################################
        df_count = df.groupby('Name')['Downtime'].count()
        df_count = df_count.reset_index()
        count_dict = dict(zip(df_count['Name'], df_count['Downtime']))

        ##############################################
        #   Sum of all DT above downtime arg
        ##############################################
        df_max = df.groupby('Name')['Downtime'].sum()
        df_max = df_max.reset_index()
        df_max.rename({'Downtime': 'DT_Sum'}, axis=1, inplace=True)
        df_max['DT_Sum'] = df_max['DT_Sum'].astype('str').str[7:]
        sum_dict = dict(zip(df_max['Name'], df_max['DT_Sum']))

        ##############################################
        #   Make dict for top 3 DTs
        ##############################################
        df_top = df[df.groupby('Name')['Downtime'].rank(ascending=False) <= 3]
        df_top.reset_index(drop=True, inplace=True)
        df_top['Tops'] = df_top['Name'] + ' ' + df_top['Time'].astype('str').str[11:] + '---' + df_top['Downtime'].astype('str').str[7:]
        del df_top['Time']
        del df_top['Downtime']

        top_list = df_top['Tops'].to_list()
        names = df.Name.unique()
        tops = {}

        for n in names:
            temp = []
            for top in top_list:
                if top.startswith(n):
                    temp.append(top.split()[1])
            tops.update({n: temp})

        ##############################################
        #   Map all dicts
        ##############################################
        dfr = df_start
        data_date = dfr['Time'].iloc[0].to_pydatetime().strftime('%A %d-%m-%Y')
        dfr.rename({'Time': 'Start'}, axis=1, inplace=True)
        dfr['Finish'] = dfr['Name'].map(finish_dict)
        dfr['Time'] = dfr['Name'].map(time_dict)
        dfr['DT_Sum'] = dfr['Name'].map(sum_dict)
        dfr['DT_Count'] = dfr['Name'].map(count_dict)
        dfr['Top_DTs'] = dfr['Name'].map(tops)

        ##############################################
        #   Reformat time cols
        ##############################################
        dfr['Start'] = dfr['Start'].astype('str').str[11:]
        dfr['Finish'] = dfr['Finish'].astype('str').str[11:]
        dfr['Time'] = dfr['Time'].astype('str').str[7:]
        dfr['Top_DTs'] = dfr['Top_DTs'].astype('str').str.replace(',', '  ')
        dfr['Top_DTs'] = dfr['Top_DTs'].astype('str').str.replace("[", '')
        dfr['Top_DTs'] = dfr['Top_DTs'].astype('str').str.replace(']', '')
        dfr['Top_DTs'] = dfr['Top_DTs'].astype('str').str.replace('"', '')
        dfr['Top_DTs'] = dfr['Top_DTs'].astype('str').str.replace("'", "")
        dfr['Top_DTs'] = dfr['Top_DTs'].astype('str').str.replace('nan', '')

        return dfr, data_date

    except Exception as e:
        print('make_dt', e)


@st.cache
def action(data):
    try:
        df_fig, name_list = make_figs(data)
        df_dt, rep_name = make_dt(data, dt_tolerance, name_list)

        ##############################################
        #   KPI
        ##############################################
        try:
            figs = df_fig['Figures'].to_list()
            tot_time = df_dt['Time'].to_list()
            minutes = []
            for t in tot_time:
                minutes.append(int(math.floor((int(t[:2]) * 60 + int(t[3:5])) / 15)) * 15)
            kpi = []
            for i, f in enumerate(figs):
                kpi.append(round((f / (minutes[i] / 60)), 2))
            df_fig['KPI'] = kpi
            df_fig = df_fig[['Name', 'Figures', 'KPI', 'Picks', 'Packs', 'Putaway', 'Moves', 'Rcv',
                             '7am', '8am', '9am', '10am', '11am', '12pm', '13pm', '14pm',
                             '15pm', '16pm', '17pm', '18pm', '19pm', '20pm', '21pm', '22pm', '23pm']]
        except Exception as e:
            print('action_kpi', e)

    except Exception as e:
        print('action_1', e)
        rep_name = 'error.xlsx'

    try:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "FIGURES"
        ws2 = wb.create_sheet(title='DOWNTIME')

        for r in dataframe_to_rows(df_fig, index=False, header=True):
            ws1.append(r)

        for r in dataframe_to_rows(df_dt, index=False, header=True):
            ws2.append(r)

        table = Table(displayName="Table1", ref="A1:" + get_column_letter(ws1.max_column) + str(ws1.max_row))
        table2 = Table(displayName="Table2", ref="A1:" + get_column_letter(ws2.max_column) + str(ws2.max_row))

        style = TableStyleInfo(name="TableStyleMedium1", showFirstColumn=True, showLastColumn=False,
                               showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        table2.tableStyleInfo = style

        ws1.column_dimensions["A"].width = 20
        ws1.column_dimensions["B"].width = 10
        ws1.column_dimensions["C"].width = 9

        ws1.column_dimensions["D"].width = 10
        ws1.column_dimensions["E"].width = 10
        ws1.column_dimensions["F"].width = 10
        ws1.column_dimensions["G"].width = 10
        ws1.column_dimensions["H"].width = 10

        ws1.column_dimensions["I"].width = 9
        ws1.column_dimensions["J"].width = 9
        ws1.column_dimensions["K"].width = 9
        ws1.column_dimensions["L"].width = 9
        ws1.column_dimensions["M"].width = 9
        ws1.column_dimensions["N"].width = 9
        ws1.column_dimensions["O"].width = 9
        ws1.column_dimensions["P"].width = 9
        ws1.column_dimensions["Q"].width = 9
        ws1.column_dimensions["R"].width = 9
        ws1.column_dimensions["S"].width = 9
        ws1.column_dimensions["T"].width = 9
        ws1.column_dimensions["U"].width = 9
        ws1.column_dimensions["V"].width = 9
        ws1.column_dimensions["W"].width = 9
        ws1.column_dimensions["X"].width = 9
        ws1.column_dimensions["Y"].width = 9

        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 12
        ws2.column_dimensions["C"].width = 12
        ws2.column_dimensions["D"].width = 12
        ws2.column_dimensions["E"].width = 12
        ws2.column_dimensions["F"].width = 12
        ws2.column_dimensions["G"].width = 60

        ws1.add_table(table)
        ws2.add_table(table2)

        xls_bytes = BytesIO(save_virtual_workbook(wb))

        try:
            zulip.Client(api_key=os.environ.get('msg_key'),
                         email=os.environ.get('msg_mail'),
                         site=os.environ.get('msg_site')).send_message({"type": "private", "to": [int(os.environ.get('msg_to'))],
                                                                        "content": f"CDReport ran at {stt('%H:%M:%S on %d-%m-%y')}"})
        except Exception as e:
            print('action_zulip', e)

        return xls_bytes, rep_name

    except Exception as e:
        print('action_2', e)


st.title("Sidi's Report")

try:
#     deta = Deta(os.environ.get('db_key'))
#     db = deta.Base(os.environ.get('db_name'))
#     access = db.get(key='access').get('value')
    access = True
except Exception as e:
    print('deta_access', e)
    access = None

if access:
    st.write('Set downtime tolerance before uploading data. Use format hh:mm:ss')
    st.write('Default value is 10 minutes')
    dt_tolerance = st.text_input(label='Downtime tolerance', label_visibility='hidden', value='00:10:00')
    st.write('Please, bear in mind that the below 5 activities are included in figures.')
    st.write('- Pack complete')
    st.write('- Pick from location')
    st.write('- Receiving')
    st.write('- Putaway putdown')
    st.write('- Move dropoff')
    uploaded_file = st.file_uploader("Choose a file", label_visibility='hidden')

    if uploaded_file is not None:
        to_dl, file_name = action(uploaded_file)

        st.download_button(
            label="Download Report",
            data=to_dl,
            file_name=file_name + '.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
else:
    st.write("I'm afraid you will have to contact your administrator.")
    st.write("...")
    st.write("Hey! Psst! The name you're trying to remember is Daniel Matyasi -> linkedin")
